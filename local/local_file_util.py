from pathlib import Path
from openpyxl import load_workbook, Workbook


def read_as_list(path):
    lines = []
    p = Path(path)
    with p.open() as f:
        line = f.readline()
        while line:
            lines.append(line.replace("\n", ""))
            line = f.readline()
    return lines


def read_as_nlist(path):
    result = []
    p = Path(path)
    with p.open() as f:
        line = f.readline()
        while line:
            line = line.replace("\n", "")
            k_v = line.split(":")
            result.append(k_v)
            line = f.readline()
    return result


def get_account():
    path = "./config/user_queue"
    acc_list = read_as_nlist(path)
    if len(acc_list) == 0:
        return None
    account = acc_list.pop()
    with open(path, "w") as w:
        for acc in acc_list:
            w.write(acc[0] + ":" + acc[1] + "\n")
    return account


def get_excel_sheet(filename, sheet):
    workbook = load_workbook(filename=filename)
    return workbook[sheet]


col = ["A", "B", "C", "D", "E", "F", "G"]


def export_excel(title, result):
    wb = Workbook()
    ws = wb.active
    sheet = wb.create_sheet(0)
    sheet.title = title
    row = 1
    for r in result:
        o = 1
        for k in r:  # 处理一行的数据
            if k == '_id':
                continue
            if k == 'account':
                co = str(col[0])
                ro = str(row)
                ws[co + ro] = r[k]
            else:
                co = str(col[o])
                ro = str(row)
                ws[co + ro] = k + ":" + r[k]
                o = o + 1
        row = row + 1  # 准备下一行
    wb.save("/home/why/Downloads/hello.xlsx")


from db.crud import CRUD

crud = CRUD("ps", "result")
db_result = crud.find({})
export_excel("hello", db_result)
